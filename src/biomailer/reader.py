from os import path
from openpyxl import workbook, load_workbook


class WorkbookReader:
    
    def __init__(self, xls_name) -> None:
        self.xls_name = xls_name
        self.wb = load_workbook(self.xls_name) # load workbook into memory
        self.ws = self.wb.active # get the active worksheet

    def transpose(self, headers=["Name", "Email", "Intended Major", "Bio"]):
        """Try to transpose worksheet arrangement into multiple rows and columns"""

        # Prepare new workbook and sheet for data
        new_wb = workbook.Workbook()
        new_ws = new_wb.active
        new_ws.append(headers) # append headers first

        record = []
        for rows in self.ws.iter_rows(min_row=1, min_col=1, max_col=1):
            for cell in rows:
                record.append(cell.value)

            # add row to new ws
            if len(record) == 4:
                new_ws.append(record)
                record.clear() # start a new record keeping

        # save new workbook with a new name
        base, ext = path.splitext(self.xls_name)
        self.xls_name = f"{base}_new{ext}"
        # reassign the self.wb and self.ws to newly created wb and ws
        new_wb.save(self.xls_name)
        self.wb, self.ws = new_wb, new_ws

    def fetch_data(self):
        """Fetches data from worksheet"""
        for row in self.ws.iter_rows(min_row=2, max_col=4, values_only=True):
            yield row


# if __name__ == '__main__':
#     a = WorkbookReader("lsl.xlsx")
#     a.fetch_data()
